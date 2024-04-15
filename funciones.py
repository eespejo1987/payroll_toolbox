#Este script contiene  funciones para modelar datos 
#que luego voy a utilizar en otros scripts

import csv
import pandas as pd
from openpyxl import load_workbook
import shutil
import os
from typing import List,Tuple

#requiere: archivo .csv 
#asegura: archivo csv como lista de listas
def leer_csv(nombre_archivo:str) -> list:
    with open(nombre_archivo,newline='', mode='r') as archivo_csv:
        csv_reader = list(csv.reader(archivo_csv, delimiter=';'))
        return csv_reader
    
#requiere: lista de listas 
#requiere: indice de inicio (ignora o no la posición[0])  
#requiere: lista de posiciones a modificar
#asegura: convierte en float las posiciones pasadas como parametro
def str_to_float(lista: list, inicio:int, posiciones:list)->list:
    i:int
    aux:str
    for x in range(inicio,len(lista),1):
        for y in range(0,len(posiciones),1):
            #i: indice posiciones
            i=posiciones[y]
            aux=(lista[x][i]).replace(',','')
            lista[x][i]=float(aux)    
    return lista

#requiere: lista de listas 
#requiere: indice de inicio (ignora o no la posición[0])  
#requiere: lista de posiciones a modificar
#asegura: si la posición esta vacia copia el valor de la fila anterior 
def rellenar_superior(lista: list, inicio:int, posiciones:list)->list:
    i:int

    for x in range(inicio,len(lista),1):
        for y in range(0,len(posiciones),1):
            i= posiciones[y]
            if lista[x][i]=='':
                lista[x][i]=lista[x-1][i]
    return lista

#requiere: lista de listas 
#requiere: indice de inicio (ignora o no la posición[0])  
#requiere: incluir encabezado
#requiere: columna a filtrar
#requiere: lista de campos a filtrar
def copiar_filtro(lista: list, inicio:int,suma_titulo:bool,filtro:int ,campos:list)->list:
    res=[]

    if suma_titulo==True:
        res.append(lista[0])
    
    for x in range(inicio,len(lista),1):
        if (lista[x][filtro] in campos)==True:
            res.append(lista[x])
    return res
    
#requiere:ruta , hoja de excel y lista de datos
#asegura: copia la información de la lista en el excel
def pegar_excel(ruta:str,hoja:str,lista:list):
    
    wb = load_workbook(ruta)
    ws = wb[hoja]
    df_datos=pd.DataFrame(lista)
    for index, row in df_datos.iterrows():
        for col_index, value in enumerate(row, start=1):
            ws.cell(row=index+1, column=col_index, value=value)
        
    wb.save(ruta)
    
#requiere:ruta origen + ruta destino + nombre archivo
#asegura: copia de archivo origen
def copiar_archivo(ruta:str, destino:str, nombre:str):
    shutil.copyfile(ruta, destino + '/' + nombre)

#requiere: lista de listas + posicion + encabezado 
#asegura: conjunto a partir de un parametro    
def conjunto_x_parametro(lista:list,posicion:int,nombre_columnas:bool)->set:

    if nombre_columnas==False:
        lista.pop(0)

    conjunto= {elemento[posicion] for elemento in lista}
    return conjunto

#requiere: conjunto(parametros) + lista + posicion a comparar
#asegura: lista de tuplas [parametro][todos los elementos que cumplen ese parametro]
#Ejemplo de uso: conjunto de contratos + nomina[contrato,cuil,nombre]
def lista_pertenencia(conjunto:set, lista:list, indice_p:int , dato:int )->list:
    conjunto_lista=list(conjunto)
    res:list=[]
    lista_aux:list=[]
    pertenece:list=[]

    for x in range(0,len(conjunto_lista),1):
        lista_aux.append(conjunto_lista[x])
        for y in range(0,len(lista),1):
            if conjunto_lista[x]==lista[y][indice_p]:
                #print(conjunto_lista[x])
                pertenece.append(lista[y][dato])
        lista_aux.append(pertenece)
        pertenece=[]
        res.append(lista_aux)
        lista_aux=[]

    return res

#requiere: lista de listas + columnas a agregar
#asegura: suma columnas al final de cada uno de los elementos de una lista
def sumar_columnas_vacias(lista:list,columnas:list)->list:
    j:int=len(lista) #ultima posicion 

    for x in range(0,len(columnas),1):
        lista[0].append(columnas[x])

    for x in range(1,len(lista),1):
        for y in range(0,len(columnas),1):
            lista[x].append('')
    return lista


#asigar por conjunto de pertenencia
#requiere: conjunto para definir si corresponde asignacion 
#requiere: posiciones para: comparar(indice) + verificar pertenencia + asignar dato  
#requiere: lista de tuplas [propiedad x][elementos que tienen propiedad x]
#Ejemplo de uso: {conjunto cuiles} + [[contrato x,[elementos x]],[contratos y,[elementos y]]
# + indice(cuil)
# una persona por contrato
def asignar_x_conjunto(asignar:set,propiedad:list,lista: list,indice:int,
                       pertenece:int, dato:int,nombre_columna:bool)->list:
    i:int=0
    for x in range(0,len(propiedad),1):
        propiedad[x][1]=set(propiedad[x][1])
    
    if nombre_columna==True:
        i=1

    for x in range(i,len(lista),1):
        if (lista[x][indice] in asignar)==True:
            lista[x][pertenece]='SI'
            for y in range(0,len(propiedad),1):
                if (lista[x][indice] in propiedad[y][1])==True:
                    lista[x][dato]=propiedad[y][0]
        else:
            lista[x][pertenece]='NO'

    return lista

#requiere: ruta de carpetas + extensión a buscar
#asegura: lista de nombres x extension
def listar_x_extension(ruta:str,extension:str)->list:
    res:list=[]
    archivos = os.listdir(ruta)

    for x in range(0,len(archivos),1):
        if archivos[x].endswith(extension):
            res.append(archivos[x])
    return res

#requiere lista de caracteres a borrar(Tupla(original,reemplazo,posicion)
#asegura: lista sin caracteres especiales
def modifica_char_esp(lista:list,char_borrar:List[Tuple[str,str,int]],nombre_columnas:bool)->list:
    i:int=0

    if nombre_columnas==True:
        i=1
    for x in range(0,len(char_borrar),1):
        j=char_borrar[x][0] #caracter a borrar/reemplazar
        k=char_borrar[x][1] #indice a borrar/reemplazar
        l=char_borrar[x][2] #posicion

        for y in range(i,len(lista),1):
            aux=(lista[y][l]).replace(j,k)
            lista[y][l]=aux
    
    return lista

def sin_acentos(palabra:str):
    minusculas_con_acentos = 'áéíóú'
    minusculas_sin_acentos = 'aeiou'
    resultado = ''

    for letra in palabra:
        if letra in minusculas_con_acentos:
            # Obtener el índice de la letra con acento
            indice = minusculas_con_acentos.index(letra)
            # Reemplazarla por la letra sin acento
            resultado += minusculas_sin_acentos[indice]
        else:
            resultado += letra

    return resultado

    
     




#pruebas
if __name__ == "__main__":
    #prueba=leer_csv('mensualxPersona3.csv')
    #prueba=str_to_float(prueba,1,[12,13])
    #prueba=rellenar_superior(prueba,1,[0,1,2])
    #resultado=(copiar_filtro(prueba,1,True,10,['1013']))
    #df_resultado=pd.DataFrame(resultado)
    #pegar_excel('Control Liq CC 023.xlsx','BBDD',resultado)
    #control_c=(lista_pertenencia(par,contratos,0,1))
    #contratos=leer_csv('contratos.csv')
    #par= (conjunto_x_parametro(contratos,0,False))
    #df_resultado.to_excel('prueb.xlsx', index=False,header=False)
    print(sin_acentos('fión'))






        

